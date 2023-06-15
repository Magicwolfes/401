#Sierra Maldonado
#!/usr/bin/env python3
from openpyxl import Workbook
from openpyxl import load_workbook
import datetime
import os

filename = 'JobHunting.xlsx'

if os.path.exists(filename):
    choose = input("The file already exists. Do you want to load the existing file? (Y/N): ")
    if choose.lower() == 'y':
        workbook = load_workbook(filename)
        sheet = workbook.active
    else:
        new_filename = input("Enter a new filename: ")
        workbook = Workbook()
        sheet = workbook.active
else:
    workbook = Workbook()
    sheet = workbook.active

# Define column names
column_names = ['Company', 'Job Name', 'Job Url Link', 'Applied Date', 'Comments', 'Last Updated']

# Write column names to the first row if the sheet is empty
if sheet.max_row == 0:
    sheet.append(column_names)

def new():
    while True:
        Company = input("Company Name: ")
        Name = input("Job Name: ")
        Url = input("Job Url Link: ")
        Applied = input("Applied date: ")
        Comments = input("Comments: ")
        Today = datetime.datetime.now().date()
        sheet.append([Company, Name, Url, Applied, Comments, Today])
        workbook.save(filename)
        choose = input("Continue or exit? ")
        if choose == 'exit':
            break
        
def update():
    while True:
        row_number = int(input("Enter the row number to update or 0 to exit: "))
        if row_number == 0:
            break
        column_number = int(input("Enter the column to update: (5 For new Comments) "))
        new_data = input("Enter new data: ")
        sheet.cell(row=row_number, column=column_number).value = new_data
        workbook.save(filename)
        
def delete_row(filename):
    workbook = load_workbook(filename)
    sheet_names = workbook.sheetnames
    print("Available sheets:")
    for index, name in enumerate(sheet_names):
        sheet = workbook[name]
        num_rows = sheet.max_row
        print(f"{index + 1}. {name} ({num_rows} rows)")
    sheet_index = int(input("Select the sheet number: ")) - 1
    sheet_name = sheet_names[sheet_index]
    row_number = int(input("Enter the row number you want to delete: "))
    sheet = workbook[sheet_name]
    row_values = []
    for cell in sheet[row_number]:
        row_values.append(cell.value)
    sheet.delete_rows(row_number)
    workbook.save(filename)
    workbook.close()
    print("Deleted row values:", row_values)
    return row_values

while True:
    choose = input("New, update, delete, or exit? ")
    if choose == "new":
        new()
    if choose == "update":
        update()
    if choose == "delete":
        delete_row(filename)
    if choose == "exit":
        break
    else:
        print("Invalid option. Please try again.")
from openpyxl import Workbook
from openpyxl import load_workbook
import datetime
import os

filename = 'JobHunting.xlsx'

if os.path.exists(filename):
    choose = input("The file already exists. Do you want to load the existing file? (Y/N): ")
    if choose.lower() == 'y':
        workbook = load_workbook(filename)
        sheet = workbook.active
    else:
        new_filename = input("Enter a new filename: ")
        workbook = Workbook()
        sheet = workbook.active
else:
    workbook = Workbook()
    sheet = workbook.active

# Define column names
column_names = ['Company', 'Job Name', 'Job Url Link', 'Applied Date', 'Comments', 'Last Updated']

# Write column names to the first row if the sheet is empty
if sheet.max_row == 0:
    sheet.append(column_names)

def new():
    while True:
        Company = input("Company Name: ")
        Name = input("Job Name: ")
        Url = input("Job Url Link: ")
        Applied = input("Applied date: ")
        Comments = input("Comments: ")
        Today = datetime.datetime.now().date()
        sheet.append([Company, Name, Url, Applied, Comments, Today])
        workbook.save(filename)
        choose = input("Continue or exit? ")
        if choose == 'exit':
            break
        
def update():
    while True:
        row_number = int(input("Enter the row number to update or 0 to exit: "))
        if row_number == 0:
            break
        column_number = int(input("Enter the column to update: (5 For new Comments) "))
        new_data = input("Enter new data: ")
        sheet.cell(row=row_number, column=column_number).value = new_data
        workbook.save(filename)
        
def delete_row(filename):
    workbook = load_workbook(filename)
    sheet_names = workbook.sheetnames
    print("Available sheets:")
    for index, name in enumerate(sheet_names):
        sheet = workbook[name]
        num_rows = sheet.max_row
        print(f"{index + 1}. {name} ({num_rows} rows)")
    sheet_index = int(input("Select the sheet number: ")) - 1
    sheet_name = sheet_names[sheet_index]
    row_number = int(input("Enter the row number you want to delete: "))
    sheet = workbook[sheet_name]
    row_values = []
    for cell in sheet[row_number]:
        row_values.append(cell.value)
    sheet.delete_rows(row_number)
    workbook.save(filename)
    workbook.close()
    print("Deleted row values:", row_values)
    return row_values

while True:
    choose = input("New, update, delete, or exit? ")
    if choose == "new":
        new()
    if choose == "update":
        update()
    if choose == "delete":
        delete_row(filename)
    if choose == "exit":
        break
    else:
        print("Invalid option. Please try again.")

